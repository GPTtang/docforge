from __future__ import annotations

from dataclasses import asdict, dataclass
from typing import Any, Iterable, Iterator, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


@dataclass
class SpreadsheetCell:
    value: str
    formula: str | None
    raw_value: str | None
    data_type: str | None
    number_format: str | None
    hyperlink: str | None
    comment: str | None
    merged: str | None


@dataclass
class SpreadsheetSheet:
    name: str
    columns: list[str]
    rows: list[list[SpreadsheetCell]]


class SpreadsheetConverter:
    def to_json(self, path: str) -> dict[str, Any]:
        values_wb = load_workbook(path, data_only=True, read_only=True, keep_links=True)
        rich_wb = load_workbook(path, data_only=False, read_only=False, keep_links=True)
        sheets: list[dict[str, Any]] = []

        try:
            for sheet in rich_wb.worksheets:
                values_sheet = values_wb[sheet.title] if sheet.title in values_wb.sheetnames else None
                sheet_payload = self._extract_sheet(sheet, values_sheet)
                sheets.append(asdict(sheet_payload))
        finally:
            values_wb.close()
            rich_wb.close()

        return {"sheets": sheets}

    def to_markdown(self, path: str) -> str:
        payload = self.to_json(path)
        parts: list[str] = []

        for sheet in payload["sheets"]:
            name = sheet["name"]
            rows = sheet["rows"]
            columns = sheet["columns"]
            parts.append(f"## Sheet: {name}")

            if not rows:
                parts.append("_(empty)_")
                parts.append("")
                continue

            table = self._rows_to_markdown_table(columns, rows)
            parts.append(table)
            parts.append("")

        return "\n".join(parts).strip() + "\n"

    def _extract_sheet(self, sheet, values_sheet=None) -> SpreadsheetSheet:
        merged_ranges = self._merged_map(sheet.merged_cells)
        value_rows: Iterator[Sequence[Any]] | None = (
            values_sheet.iter_rows(values_only=True) if values_sheet else None
        )
        rows: list[list[SpreadsheetCell]] = []

        for row in sheet.iter_rows():
            value_row = next(value_rows, None) if value_rows else None
            cells: list[SpreadsheetCell] = []

            for col_idx, cell in enumerate(row, start=1):
                display_value = self._resolve_display_value(value_row, col_idx, cell)
                cell_payload = SpreadsheetCell(
                    value=display_value,
                    formula=self._extract_formula(cell),
                    raw_value=self._normalize_cell(cell.value),
                    data_type=cell.data_type,
                    number_format=getattr(cell, "number_format", None),
                    hyperlink=getattr(cell.hyperlink, "target", None),
                    comment=self._normalize_cell(getattr(getattr(cell, "comment", None), "text", None)),
                    merged=merged_ranges.get(cell.coordinate),
                )
                cells.append(cell_payload)

            trimmed = self._trim_trailing_empty_cells(cells)
            if trimmed:
                rows.append(trimmed)

        columns = self._build_column_labels(max((len(row) for row in rows), default=0))
        return SpreadsheetSheet(name=sheet.title, columns=columns, rows=rows)

    def _rows_to_markdown_table(self, columns: list[str], rows: list[list[dict]]) -> str:
        width = max(len(columns), max(len(r) for r in rows)) if rows else len(columns)
        padded_rows = [row + [{} for _ in range(max(0, width - len(row)))] for row in rows]

        header = [self._render_cell(cell) for cell in padded_rows[0]] if padded_rows else []
        if not any(header):
            # First row is empty: substitute column letters as the header and skip that row.
            header = [self._escape_md(col or f"Column {i + 1}") for i, col in enumerate(columns or ["" for _ in range(width)])]
        body = padded_rows[1:] if len(padded_rows) > 1 else []

        lines = ["| " + " | ".join(header) + " |", "| " + " | ".join(["---"] * width) + " |"]

        for row in body:
            rendered = [self._render_cell(cell) for cell in row]
            lines.append("| " + " | ".join(rendered) + " |")

        return "\n".join(lines)

    def _render_cell(self, cell: dict | None) -> str:
        if not cell:
            return ""

        value = self._escape_md(cell.get("value", "") or "")
        formula = cell.get("formula")
        hyperlink = cell.get("hyperlink")
        comment = cell.get("comment")
        merged = cell.get("merged")

        parts: list[str] = []

        if formula and (not value or value == formula):
            text = f"`{formula}`"
        else:
            text = value
        if hyperlink:
            text = f"[{value or hyperlink}]({hyperlink})"
        parts.append(text)

        if formula and formula != value:
            parts.append(f"`{formula}`")

        if comment:
            parts.append(f"<!-- {self._escape_md(comment)} -->")

        if merged:
            parts.append(f"`{merged}`")

        return " ".join(part for part in parts if part).strip()

    def _build_column_labels(self, count: int) -> list[str]:
        return [get_column_letter(idx + 1) for idx in range(count)]

    def _merged_map(self, merged_ranges: Iterable) -> dict[str, str]:
        mapping: dict[str, str] = {}
        for cell_range in getattr(merged_ranges, "ranges", []):
            min_col, min_row, max_col, max_row = cell_range.bounds
            coord = cell_range.coord
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    mapping[f"{get_column_letter(col)}{row}"] = coord
        return mapping

    def _resolve_display_value(self, value_row: Sequence[Any] | None, col_idx: int, cell) -> str:
        if value_row and col_idx - 1 < len(value_row):
            cached_value = value_row[col_idx - 1]
            if cached_value is not None:
                return self._normalize_cell(cached_value)

        return self._normalize_cell(cell.value)

    def _extract_formula(self, cell) -> str | None:
        if cell.data_type == "f" and cell.value:
            text = str(cell.value)
            return text if text.startswith("=") else f"={text}"
        return None

    def _cell_has_payload(self, cell: SpreadsheetCell) -> bool:
        return any([cell.value, cell.formula, cell.hyperlink, cell.comment, cell.merged])

    def _trim_trailing_empty_cells(self, cells: list[SpreadsheetCell]) -> list[SpreadsheetCell]:
        trimmed = list(cells)
        while trimmed and not self._cell_has_payload(trimmed[-1]):
            trimmed.pop()
        return trimmed

    def _normalize_cell(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value)

    def _escape_md(self, text: str) -> str:
        return text.replace("|", "\\|").replace("\n", " ").replace("\r", " ").strip()
